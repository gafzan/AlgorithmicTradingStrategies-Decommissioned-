"""
excel_tools.py
"""
import logging
from pathlib import Path

import pandas as pd
import xlrd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Color
from openpyxl.styles.borders import Border, Side
from openpyxl.utils import get_column_letter

# for main module
from tkinter import *
from tkinter import filedialog
from financial_analysis.finance_tools import return_and_risk_analysis
from database.config_database import __BASE_FOLDER__, __BACK_TEST_FOLDER__

# logger
logger = logging.getLogger(__name__)
logger.setLevel(logging.INFO)
formatter = logging.Formatter('%(asctime)s : %(module)s : %(funcName)s : %(message)s')
stream_handler = logging.StreamHandler()
stream_handler.setFormatter(formatter)
logger.addHandler(stream_handler)

__COLUMN_NAMES_FONT__ = Font(bold=True,
                             italic=False,
                             underline='none',
                             strike=False,
                             color='FFFFFF')

__COLUMN_NAMES_FILL__ = PatternFill(patternType='solid', fgColor=Color(rgb='00B050'))
__DATE_COLUMN_FILL__ = PatternFill(patternType='solid', fgColor=Color(rgb='EAEAEA'))

__BORDER__ = Border(left=Side(border_style='thin', color='000000'), right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='000000'), bottom=Side(border_style='thin', color='000000'))


def save_df(df_list: {list, pd.DataFrame}, workbook_name: str = None, folder_path: str = None, full_path: str = None,
            sheet_name_list: {str, list}=None):
    """Assumes df_list is a list of DataFrame, name is a string and sheet_name_list is a list of strings
    Saves all DataFrames in df_list to a sheet with the corresponding name in sheet_list"""

    _check_save_load_input(workbook_name, folder_path, full_path)  # check the inputs

    if full_path is None:
        file_name = '\\' + workbook_name + '.xlsx'
        full_path = folder_path + file_name
    writer = pd.ExcelWriter(full_path, engine='xlsxwriter')

    if sheet_name_list is None:
        if type(df_list) == list:
            sheet_name_list = []
            for i in range(len(df_list)):
                sheet_name_list.append(f'Sheet{i + 1}')
        else:
            sheet_name_list = 'Sheet1'

    if type(df_list) == type(sheet_name_list) == list:
        if len(df_list) == len(sheet_name_list):
            for i in range(len(df_list)):
                if len(sheet_name_list[i]) > 31:
                    logger.warning(f"'{sheet_name_list[i]}' is too long (needs to be less than 31 characters). "
                                   f"\n'{sheet_name_list[i]}' will be shortened to '{sheet_name_list[i][:31]}'.")
                df_list[i].to_excel(writer, sheet_name=sheet_name_list[i][:31])  # write the DataFrame into excel
        else:
            raise ValueError(r'"df_list" and "sheet_name_list" are not of the same length')
    elif type(df_list) == list or type(sheet_name_list) == list:
        raise ValueError(r'"df_list" and "sheet_name_list" are not of the same type')
    else:
        if len(sheet_name_list) > 31:
            logger.warning(f"'{sheet_name_list}' is too long (needs to be less than 31 characters). "
                           f"\n'{sheet_name_list}' will be shortened to {sheet_name_list[:31]}")
        df_list.to_excel(writer, sheet_name=sheet_name_list[:31])  # write the DataFrame into excel
    writer.save()  # close the Pandas Excel writer and output the excel file


def load_df(workbook_name: str = None, folder_path: str = None, full_path: str = None, sheet_name: str = 'Sheet1',
            first_column_index=True, sheet_name_error_handling: bool = True):
    """Assumes path_name and file_name are strings
    Returns a DataFrame"""

    _check_save_load_input(workbook_name, folder_path, full_path)  # check the inputs

    if full_path is None:
        base_path = Path(folder_path)
        excel_file_name = workbook_name + '.xlsx'
        xl_file_fullname = base_path / excel_file_name
    else:
        xl_file_fullname = Path(full_path)

    if not xl_file_fullname.exists():
        print('No such file or directory!')
        sys.exit(-1)

    if sheet_name_error_handling:
        try:
            data = pd.read_excel(xl_file_fullname, sheet_name=sheet_name)
        except xlrd.biffh.XLRDError:
            # if the sheet name does not exist, ask the user to name one
            workbook = load_workbook(xl_file_fullname)
            if len(workbook.sheetnames) == 1:
                logger.warning(
                    "'{}' does not exist. Loading '{}' sheet since it is the only sheet in the workbook.".format(
                        sheet_name, workbook.sheetnames[0]))
                data = pd.read_excel(xl_file_fullname, sheet_name=workbook.sheetnames[0])
            else:
                print(f"'{sheet_name}' does not exist. Chose one sheet from below:")
                for sht_name in workbook.sheetnames:
                    print(f"{workbook.sheetnames.index(sht_name) + 1}: {sht_name}")
                ask_user = True
                while ask_user:
                    try:
                        number = int(input(f"Input a number between 1 and {len(workbook.sheetnames)}: "))
                        if number < 0:
                            raise ValueError
                        sheet_name = workbook.sheetnames[number - 1]
                        ask_user = False
                    except (ValueError, IndexError):  # catch all errors
                        pass
                data = pd.read_excel(xl_file_fullname, sheet_name=sheet_name)
    else:
        data = pd.read_excel(xl_file_fullname, sheet_name=sheet_name)
    if first_column_index:
        data.set_index(list(data)[0], inplace=True, drop=True)
    return data


def load_df_from_chosen_excel_file(sheet_name='Sheet1', first_column_index: bool = True, initialdir: str = __BASE_FOLDER__,
                                   title_msg: str = 'Select excel file'):
    # chose excel file to format
    Tk().withdraw()
    excel_file_path = filedialog.askopenfile(initialdir=initialdir, title=title_msg).name
    return load_df(full_path=excel_file_path, sheet_name=sheet_name, first_column_index=first_column_index)


def _check_save_load_input(workbook_name: str = None, folder_path: str = None, full_path: str = None)-> None:
    """
    Checks so only full_path or both workbook_name and folder_path is provided.
    :param workbook_name: string
    :param folder_path: string
    :param full_path: string with full path including the file type (e.g. xlsx)
    :return:
    """
    if full_path is None and not (workbook_name is not None and folder_path is not None):
        raise ValueError("Please provide both workbook_name and folder_path.")
    elif full_path is not None and (workbook_name is not None or folder_path is not None):
        raise ValueError("Please only provide full_path.")
    elif full_path is None and workbook_name is None and folder_path is None:
        raise ValueError("Please only provide full_path or both workbook_name and folder_path.")
    return


def excel_files_in_directory(directory: {Path, str}) -> list:
    """Assumes that directory is a string showing the directory path. Returns a list of strings of the name of all files
    ending with '.xlsx'"""
    if isinstance(directory, str):
        directory = Path(directory)
    return [f.name for f in directory.glob("*.xlsx")]


def choose_excel_file_from_folder(folder_path: str) -> str:
    """Assumes folder_path is a string. Returns a string of the excel file chosen by the user from the folder"""
    # print the available excel files
    counter = 0
    excel_name_list = excel_files_in_directory(folder_path)
    for excel_name in excel_name_list:
        counter += 1
        print('{}: {}'.format(counter, excel_name))

    # chose the excel file
    ask_user = True
    number = 1
    while ask_user:
        try:
            number = int(input('Enter a number between 1 and {}:'.format(counter)))
            assert 1 <= number <= counter
            ask_user = False
        except (ValueError, AssertionError):
            pass
    return excel_name_list[number - 1]


def format_requested_data_workbook(complete_workbook_path: str):
    """
    Formats the workbook that is the result of a data request from the financial database
    :param complete_workbook_path: destination of the workbook
    :return:
    """

    workbook = load_workbook(complete_workbook_path)

    for sheet in [workbook[sheet_name] for sheet_name in workbook.sheetnames if sheet_name != 'underlying_data']:
        sheet.cell(row=1, column=1).value = 'Date'
        for col in range(1, sheet.max_column + 1):
            if col == 1:  # first column has the dates
                sheet.column_dimensions[get_column_letter(col)].width = 15
            else:
                sheet.column_dimensions[get_column_letter(col)].width = 18
            basic_formatting(sheet=sheet, col=col, cell_formatting='#,##0.00')
        sheet.freeze_panes = 'B2'

    try:
        underlying_data_sheet = workbook['underlying_data']
    except KeyError:
        workbook.save(complete_workbook_path)
        return
    underlying_data_sheet.cell(row=1, column=1).value = 'ticker'
    for col in range(1, underlying_data_sheet.max_column + 1):
        if col == 1:
            underlying_data_sheet.column_dimensions[get_column_letter(col)].width = 10
        else:
            underlying_data_sheet.column_dimensions[get_column_letter(col)].width = 28
        for row in range(1, underlying_data_sheet.max_row + 1):
            cell = underlying_data_sheet.cell(row=row, column=col)
            if row == 1:
                cell.font = __COLUMN_NAMES_FONT__
                cell.fill = __COLUMN_NAMES_FILL__
            elif col == 1:
                cell.fill = __DATE_COLUMN_FILL__
            cell.alignment = Alignment(horizontal='left', vertical='center')
            cell.border = __BORDER__

    workbook.save(complete_workbook_path)


def format_risk_return_analysis_workbook(complete_workbook_path: str):
    """Assumes that complete_workbook_path is a string containing the path to the Excel workbook. Changes the format
    of certain sheets."""
    workbook = load_workbook(complete_workbook_path)

    for sheet_name in workbook.sheetnames:
        if sheet_name == 'Risk and return':
            risk_return_sheet = workbook[sheet_name]
            risk_return_sheet.column_dimensions['A'].width = 19
            for col in range(1, risk_return_sheet.max_column + 1):
                if col == 1:
                    risk_return_sheet.column_dimensions[get_column_letter(col)].width = 19
                else:
                    risk_return_sheet.column_dimensions[get_column_letter(col)].width = 12
                for row in range(1, risk_return_sheet.max_row + 1):
                    cell = risk_return_sheet.cell(row=row, column=col)
                    if row == 1:
                        cell.font = __COLUMN_NAMES_FONT__
                        cell.fill = __COLUMN_NAMES_FILL__
                    elif col == 1:
                        cell.fill = __DATE_COLUMN_FILL__
                    elif row == 4:
                        cell.number_format = '#,##0.00'
                    else:
                        cell.number_format = '0.00%'
                    if col == 1:
                        cell.alignment = Alignment(horizontal='left', vertical='center')
                    else:
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = __BORDER__
        elif sheet_name == 'Description':
            sheet = workbook[sheet_name]
            for col in range(1, sheet.max_column + 1):
                sheet.column_dimensions[get_column_letter(col)].width = 30
                basic_formatting(sheet=sheet, col=col, horizontal_alignment='left')
        elif sheet_name in ['Rolling 1Y return', 'Rolling 1Y volatility', 'Rolling 1Y drawdown']:
            # format rolling sheets
            rolling_sheet = workbook[sheet_name]
            rolling_sheet.cell(row=1, column=1).value = 'Date'
            for col in range(1, rolling_sheet.max_column + 1):
                rolling_sheet.column_dimensions[get_column_letter(col)].width = 12
                basic_formatting(rolling_sheet, col)
            rolling_sheet.freeze_panes = 'B2'
        elif 'Return table' in sheet_name:
            return_table_sheet = workbook[sheet_name]
            return_table_sheet.column_dimensions[get_column_letter(return_table_sheet.max_column)].width = 12
            for col in range(1, return_table_sheet.max_column + 1):
                basic_formatting(sheet=return_table_sheet, col=col, last_column_bold=True)
        else:
            sheet = workbook[sheet_name]
            sheet.cell(row=1, column=1).value = 'Date'
            if sheet_name == 'Performance':
                cell_formatting = '#,##0.00'
            else:
                cell_formatting = '0.00%'
            for col in range(1, sheet.max_column + 1):
                sheet.column_dimensions[get_column_letter(col)].width = 12
                basic_formatting(sheet=sheet, col=col, cell_formatting=cell_formatting)
            sheet.freeze_panes = 'B2'
    workbook.save(complete_workbook_path)


def basic_formatting(sheet, col, cell_formatting='0.00%', last_column_bold=False, horizontal_alignment: str = 'center'):
    """
    :param last_column_bold:
    :param sheet:
    :param col:
    :param cell_formatting:
    """
    for row in range(1, sheet.max_row + 1):
        cell = sheet.cell(row=row, column=col)
        if row == 1:
            cell.font = __COLUMN_NAMES_FONT__
            cell.fill = __COLUMN_NAMES_FILL__
        elif col == 1:
            cell.number_format = 'D MMM YYYY'
            cell.fill = __DATE_COLUMN_FILL__
        else:
            cell.number_format = cell_formatting
        cell.alignment = Alignment(horizontal=horizontal_alignment, vertical='center')
        cell.border = __BORDER__
        if last_column_bold and col == sheet.max_column and row != 1:
            cell.fill = PatternFill(patternType='solid', fgColor=Color(rgb='EBF1DE'))
            cell.font = Font(bold=True)


def format_excel_workbook_index_time_series():
    # get the underlying data
    df = load_df_from_chosen_excel_file()
    sheet_name_df_dict = return_and_risk_analysis(df, print_results=True, normalize=False)

    # save the result
    save_file_path = filedialog.asksaveasfile(initialdir=__BACK_TEST_FOLDER__, title='Save excel file',
                                              filetypes=(("Excel files", "*.xlsx"), ("All files", "*.*"))).name
    save_file_path += '.xlsx'
    print(save_file_path)
    logger.info('Saving results to excel: {}'.format(save_file_path))
    save_df(list(sheet_name_df_dict.values()), full_path=save_file_path, sheet_name_list=list(sheet_name_df_dict.keys()))
    logger.info('Formatting excel file...')
    format_risk_return_analysis_workbook(save_file_path)
    logger.info('Done!')


def main():
    # chose excel file to format
    Tk().withdraw()
    excel_file_path = filedialog.askopenfile(initialdir=__BASE_FOLDER__, title='Select excel file to format').name

    # get the underlying data
    format_requested_data_workbook(excel_file_path)


if __name__ == '__main__':
    main()

